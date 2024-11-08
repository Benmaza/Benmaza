import openpyxl

# Pide al usuario que ingrese el nombre del archivo de Excel
nombre_archivo = input("Por favor, ingresa el nombre del archivo de Excel que deseas abrir: ")

# Abre el archivo de Excel
wb = openpyxl.load_workbook(nombre_archivo)

# Crea una nueva hoja
new_sheet = wb.create_sheet('Tendencia')

# Itera sobre todas las hojas en el libro
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    
    # Toma el rango de celdas que quieres
    cells = sheet['E19:E24']
    
    # Copia los datos a la nueva hoja
    for i, cell in enumerate(cells, start=1):
        new_sheet[sheet_name + 'A' + str(i)] = cell.value

# Guarda el archivo de Excel
wb.save(nombre_archivo)